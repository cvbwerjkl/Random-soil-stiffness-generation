import numpy as np
from scipy import stats
from Graphics.total import total

"""
from formStiffnessBernoulliBeam import formStiffnessBernoulliBeam
from formStiffnessSpring import formStiffnessSpring
from random_gen import random_gen
"""
from sumVarArray import sumVarArray
from solution import solution

if __name__ == "__main__":
    print("Task #1")
    # !!!!Input parameters constants

    P = -500  # distributed load
    # Beam stiffness
    E = 36e4
    t = 40
    I = (t ** 3) / 12
    EI = E * I
    numberSprings = 300
    generations = 1000  # number of random generation

    """
    # !!!!Input parameters variables
    springStiffnessMean = 2e5  # mean spring stiffness
    deviation = 0.3  # spring stiffness standard deviation
    Lbeam = 40
    corRadius = 20  # correlation radius
    """

    # !!!Input parameters variables arrays

    """
    deviation_array = np.array([0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4])
    springStiffnessMean_array = np.array([1e5, 1.5e5, 2e5, 2.5e5, 3e5])
    Lbeamarray = np.array([20, 25, 30, 35, 40])
    corRadius_array = np.array([0.05, 0.1, 0.25, 0.5, 1, 2, 5, 10, 15, 20])
    """

    deviation_array = np.array([0.05, 0.1, 0.15, 0.2, 0.25, 0.3])
    springStiffnessMean_array = np.array([1e5, 1.5e5, 2e5, 2.5e5, 3e5])
    Lbeamarray = np.array([20, 25, 30, 35, 40])
    corRadius_array = np.array([0.05, 0.1, 0.25, 0.5, 1, 2, 5, 10, 15, 20])

    # summary variables array creation
    sumArray = sumVarArray(deviation_array, springStiffnessMean_array, Lbeamarray, corRadius_array)

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    exTitl = 'B' + str(1)
    ws[exTitl] = 'maxTilt'
    exTitl = 'C' + str(1)
    ws[exTitl] = 'minTilt'
    exTitl = 'D' + str(1)
    ws[exTitl] = 'meanTilt'
    exTitl = 'E' + str(1)
    ws[exTitl] = 'sd'
    exTitl = 'F' + str(1)
    ws[exTitl] = 'tilt_s95'
    exTitl = 'G' + str(1)
    ws[exTitl] = 'tilt_s99'
    exTitl = 'H' + str(1)
    ws[exTitl] = 'error_095'
    exTitl = 'I' + str(1)
    ws[exTitl] = 'error_0.99'
    exTitl = 'J' + str(1)
    ws[exTitl] = 'meanSettlement'

    pointVal = np.empty(sumArray.size)
    pointTilt = np.empty(sumArray.size)

    cell_N = 2
    with open("output.txt", "w") as f:
        for i in sumArray:
            deviation = i[0]
            springStiffnessMean = i[1]
            Lbeam = i[2]
            corRadius = i[3]
            P_l = (P / Lbeam) * 40  # distibuted load with Lbeam dependance


            #  !!!Solution
            sAvg, tilt = solution(Lbeam, EI, P_l, numberSprings, corRadius, springStiffnessMean, deviation, generations)

            # Tilt data evaluation

            sd = stats.tstd(tilt)  # standard deviation (n-1)
            meanT = stats.tmean(tilt)  # mean
            maxT = stats.tmax(tilt)  # max
            minT = stats.tmin(tilt)  # min

            confidence1 = 0.95  # confidence 0.95
            tilt_s95 = meanT * stats.t.ppf((1 - (1 - confidence1) / 2), generations - 1)
            confidence2 = 0.99  # confidence 0.99
            tilt_s99 = meanT * stats.t.ppf((1 - (1 - confidence2) / 2), generations - 1)

            # print(min1)
            # print(mean)
            e1 = float(tilt_s95) * (sd / (generations) ** 0.5)
            e2 = float(tilt_s99) * (sd / (generations) ** 0.5)

            # savg data evaluation
            meanS = stats.tmean(sAvg)  # mean

            pointVal[cell_N - 2] = (deviation * meanS * (corRadius ** 0.33333)) / Lbeam
            pointTilt[cell_N - 2] = tilt_s95

            ss = str((deviation * meanS * (corRadius ** 0.33333)) / Lbeam) + ' ' + str(tilt_s95) + ' ' + str(deviation) + ' ' + str(meanS) + ' ' + str(corRadius) + ' ' + str(Lbeam)
            print(ss, file=f)


            # !!!Save to xslx
            exTitl = 'A' + str(cell_N)
            ws[exTitl] = 'L_' + str(Lbeam) + '_S_' + str(int(meanS * 1000)) + '_SD_' + str(deviation) + '_Rc_' + str(corRadius)
            exTitl = 'B' + str(cell_N)
            ws[exTitl] = maxT
            exTitl = 'C' + str(cell_N)
            ws[exTitl] = minT
            exTitl = 'D' + str(cell_N)
            ws[exTitl] = meanT
            exTitl = 'E' + str(cell_N)
            ws[exTitl] = sd
            exTitl = 'F' + str(cell_N)
            ws[exTitl] = tilt_s95
            exTitl = 'G' + str(cell_N)
            ws[exTitl] = tilt_s99
            exTitl = 'H' + str(cell_N)
            ws[exTitl] = e1
            exTitl = 'I' + str(cell_N)
            ws[exTitl] = e2
            exTitl = 'J' + str(cell_N)
            ws[exTitl] = meanS

            print(meanT, " ", meanS)

            cell_N += 1

    wb.save("summary.xlsx")

    total(pointVal, pointTilt)
